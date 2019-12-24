from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1']=42
wb.save("zhc1.xlsx")#D:\专业软件\Python\Project_ZHC\GH       project路径
print(ws['A1'])
